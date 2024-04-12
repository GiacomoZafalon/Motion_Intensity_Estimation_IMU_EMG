import torch
import openpyxl

imu_data = openpyxl.load_workbook('c:/Users/giaco/OneDrive/Desktop/Università/Tesi_Master/Dataset/P1/W1/imu/imu_p1_a1_1.xlsx')
sheet_names = imu_data.sheetnames

num_sensors = 3

sheets = {}
imu_p1_a1_1_data = []
imu_p1_a1_1_all = []
imu_p1_a1_1_tot = []
imu_p1_a1_1_fin = []
# Load sheets based on the number of sensors
for i in range(1, num_sensors + 1):
    imu_p1_a1_1_data = []
    sheet_name = f"Foglio{i}"
    if sheet_name in sheet_names:
        sheets[sheet_name] = imu_data[sheet_name]
        for row in sheets[sheet_name].iter_rows(values_only=True):
            imu_p1_a1_1_data.append(row[:])
        imu_p1_a1_1_all.append(imu_p1_a1_1_data)
imu_p1_a1_1 = torch.tensor(imu_p1_a1_1_all)

for i in range(imu_p1_a1_1.shape[1]):
    imu_p1_a1_1_tot = []
    for j in range(imu_p1_a1_1.shape[0]):
        imu_p1_a1_1_tot.append(imu_p1_a1_1[j][i].tolist())
    imu_p1_a1_1_fin.append(imu_p1_a1_1_tot)

data_p1_a1_1 = torch.tensor(imu_p1_a1_1_fin)

emg_data = openpyxl.load_workbook('c:/Users/giaco/OneDrive/Desktop/Università/Tesi_Master/Dataset/P1/W1/emg/emg_p1_a1_1.xlsx')
emg_data = emg_data.active
# Access values from the first column
MVC = 500
emg_p1_a1_1 = []
for col in emg_data.iter_cols(values_only=True):
    emg_p1_a1_1.append(col[:])
    max_each_sensor, index_max_each_sensor = torch.max(torch.tensor(emg_p1_a1_1), axis=1)
max_tot = max_each_sensor.sum()
label_p1_a1_1 = max_tot / MVC

p1_a1_1 = data_p1_a1_1, label_p1_a1_1
# print(p1_a1_1)

class DataProcessor:
    def __init__(self, person, action, attempt, num_sensors=3, mvc=500):
        self.person = person
        self.action = action
        self.attempt = attempt
        self.num_sensors = num_sensors
        self.mvc = mvc

    def load_imu_data(self, person, action, attempt):
        imu_path = f"c:/Users/giaco/OneDrive/Desktop/Università/Tesi_Master/Dataset/P{person}/W{action}/imu/imu_p{person}_a{action}_{attempt}.xlsx"
        # print('aaaaa', person, action, attempt)
        imu_data = openpyxl.load_workbook(imu_path)
        # print(imu_path)
        sheet_names = imu_data.sheetnames
        # imu_all = []
        
        sheets = {}
        imu_p1_a1_1_data = []
        imu_p1_a1_1_all = []
        imu_p1_a1_1_tot = []
        imu_p1_a1_1_fin = []
        # Load sheets based on the number of sensors
        for i in range(1, num_sensors + 1):
            imu_p1_a1_1_data = []
            sheet_name = f"Foglio{i}"
            if sheet_name in sheet_names:
                sheets[sheet_name] = imu_data[sheet_name]
                for row in sheets[sheet_name].iter_rows(values_only=True):
                    imu_p1_a1_1_data.append(row[:])
                imu_p1_a1_1_all.append(imu_p1_a1_1_data)
        imu_p1_a1_1 = torch.tensor(imu_p1_a1_1_all)

        for i in range(imu_p1_a1_1.shape[1]):
            imu_p1_a1_1_tot = []
            for j in range(imu_p1_a1_1.shape[0]):
                imu_p1_a1_1_tot.append(imu_p1_a1_1[j][i].tolist())
            imu_p1_a1_1_fin.append(imu_p1_a1_1_tot)

        imu_tensor = torch.tensor(imu_p1_a1_1_fin)
        
        return imu_tensor

    def load_emg_data(self, person, action, attempt):
        emg_path = f"c:/Users/giaco/OneDrive/Desktop/Università/Tesi_Master/Dataset/P{person}/W{action}/emg/emg_p{person}_a{action}_{attempt}.xlsx"
        emg_data = openpyxl.load_workbook(emg_path)
        emg_data = emg_data.active
        emg_all = []

        for col in emg_data.iter_cols(values_only=True):
            emg_all.append(col[:])

        emg_tensor = torch.tensor(emg_all)
        return emg_tensor

    def process_data(self):
        all_data = []

        for person in range(1, self.person + 1):
            for action in range(1, self.action + 1):
                for attempt in range(1, self.attempt + 1):
                    imu_data = self.load_imu_data(person, action, attempt)
                    # print(imu_data)
                    emg_data = self.load_emg_data(person, action, attempt)

                    max_each_sensor, _ = torch.max(emg_data, axis=1)
                    max_tot = max_each_sensor.sum()
                    label = max_tot / self.mvc

                    all_data.append((imu_data, label))

        return all_data

# Example usage:
person = 1
action = 1
attempt = 2

processor = DataProcessor(person, action, attempt)
all_data = processor.process_data()
imu_data, label = all_data
print(all_data)