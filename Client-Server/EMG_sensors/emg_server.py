import os
import csv
from openpyxl import load_workbook
import sys

# Ensure the path to your module is correct
sys.path.append('C:/Users/giaco/OneDrive/Desktop/Università/Tesi_Master/GitHub/Client-Server/IMU_sensors')
import save_dir_info

# Load variables from save_dir_info module
person = save_dir_info.person
weight = save_dir_info.weight
attempt = save_dir_info.attempt

# Path to Excel and CSV files using os.path.join for better cross-platform support
base_dir = r'C:\Users\giaco\OneDrive\Desktop\Università\Tesi_Master\GitHub'
excel_filename = os.path.join(base_dir, 'emg_data.xlsx')
csv_dir = os.path.join(base_dir, f'Dataset\P{person}\W{weight}\A{attempt}\emg')
os.makedirs(csv_dir, exist_ok=True)
csv_filename = os.path.join(csv_dir, 'emg_data.csv')

# Load Excel workbook
wb = load_workbook(excel_filename)

# Access specific worksheet
ws = wb['Data In']

# Open CSV file for writing
with open(csv_filename, 'w', newline='') as csvfile:
    csvwriter = csv.writer(csvfile)

    # Write headers if needed
    # csvwriter.writerow(['Column1', 'Column2', 'Column3'])

    # Iterate over rows and columns to access data
    for row in ws.iter_rows(min_row=8, values_only=True):  # Data starts from row 8
        # Check if all values in the row are None
        if all(value is None for value in row):
            break  # Stop iterating if all values in the row are None
        # Extract values from the second, third, and fourth columns (columns B, C, D in Excel)
        column_values = row[1:4]
        csvwriter.writerow(column_values)
