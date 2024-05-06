from openpyxl import load_workbook
import csv
import sys
sys.path.append('C:/Users/giaco/OneDrive/Desktop/Università/Tesi_Master/GitHub/Client-Server/IMU_sensors')
import save_dir_info

person = save_dir_info.person
weight = save_dir_info.weight
attempt = save_dir_info.attempt

# Load Excel workbook
wb = load_workbook(rf'c:\Users\giaco\OneDrive\Desktop\Università\Tesi_Master\GitHub\Dataset\emg_data.xlsx')

# Access specific worksheet
ws = wb['Data In']

# Create CSV file name based on person, weight, and attempt
csv_filename = rf'c:\Users\giaco\OneDrive\Desktop\Università\Tesi_Master\GitHub\Dataset\P{person}\W{weight}\A{attempt}\emg\emg_data.csv'

# Open CSV file for writing
with open(csv_filename, 'w', newline='') as csvfile:
    csvwriter = csv.writer(csvfile)
    
    # Iterate over rows and columns to access data
    for row in ws.iter_rows(min_row=8, values_only=True):  # Data starts from row 8
        # Check if all values in the row are None
        if all(value is None for value in row):
            break  # Stop iterating if all values in the row are None
        # Extract values from the second, third, and fourth columns
        column_values = row[1:4]
        csvwriter.writerow(column_values)
